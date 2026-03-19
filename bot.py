"""
契約代行アシスタント Telegram Bot
機能:
1. 名刺の自動作成（法人一覧から選択してPNG生成）
2. 支払い依頼フォーム（銀行一覧選択・検索、スプレッドシートに記録）
3. 稼働データ入力フォーム（稼働者名・店舗数・日付・iPhone機種・容量・台数）
"""

import io
import json
import logging
import os
import subprocess
from datetime import datetime, timezone, timedelta

import requests
import openpyxl
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    BotCommand,
)
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)

from banks_data import MAJOR_BANKS, search_banks
from meishi_generator import create_business_card

# ── Configuration ──────────────────────────────────────────────────────────
BOT_TOKEN = os.environ.get("BOT_TOKEN", "8789721641:AAFKm0JIBMZKcIhqc6htSgnwl3fvTj2PY2c")
SPREADSHEET_ID = os.environ.get("SPREADSHEET_ID", "104JfX8b4VuE6T2yGKI6hLL58z3gZSKQ339TLnQ_Y2iI")
TRANSFER_NOTIFY_GROUP_ID = int(os.environ.get("TRANSFER_NOTIFY_GROUP_ID", "-5006222520"))  # 振込依頼通知先グループ
# 名刺作成機能の許可ユーザーネームリスト（@なしのユーザーネームで指定）
MEISHI_ALLOWED_USERS: set[str] = set(
    u.strip().lstrip("@").lower()
    for u in os.environ.get("MEISHI_ALLOWED_USERS", "kk_12345,ks19970606").split(",")
    if u.strip()
)
RCLONE_CONFIG = os.environ.get("RCLONE_CONFIG", "/home/ubuntu/.gdrive-rclone.ini")
GDRIVE_ACCESS_TOKEN = os.environ.get("GDRIVE_ACCESS_TOKEN", "")
JST = timezone(timedelta(hours=9))

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Conversation states ────────────────────────────────────────────────────
# 名刺作成
MEISHI_SELECT = 100
MEISHI_PAGE = 101

# 支払い依頼
TRANSFER_NAME = 200
TRANSFER_BANK = 201
TRANSFER_BANK_SEARCH = 202
TRANSFER_BRANCH = 203
TRANSFER_TYPE = 204
TRANSFER_ACCOUNT = 205
TRANSFER_AMOUNT = 206
TRANSFER_CONFIRM = 207

# 稼働報告
REPORT_NAME = 300
REPORT_SHOP = 301
REPORT_DATE = 302
REPORT_DATE_INPUT = 303
REPORT_MODEL = 304
REPORT_CAPACITY = 305
REPORT_QUANTITY = 306
REPORT_ADD_MORE = 307
REPORT_CONFIRM = 308

# 代行登録
REG_NAME = 400
REG_PHONE = 401
REG_ADDRESS = 402
REG_BANK = 403
REG_BANK_SEARCH = 404
REG_BRANCH = 405
REG_ACCOUNT = 406
REG_CONFIRM = 407


# ── Google Drive helpers ───────────────────────────────────────────────────
def get_access_token() -> str:
    """Google Drive APIのアクセストークンを取得する"""
    # 1. GDRIVE_ACCESS_TOKEN 環境変数（直接トークン文字列）
    if GDRIVE_ACCESS_TOKEN:
        return GDRIVE_ACCESS_TOKEN
    # 2. GDRIVE_TOKEN 環境変数（JSON形式のトークン）
    gdrive_token_json = os.environ.get("GDRIVE_TOKEN", "")
    if gdrive_token_json:
        try:
            token_data = json.loads(gdrive_token_json)
            return token_data.get("access_token", "")
        except (json.JSONDecodeError, KeyError):
            pass
    # 3. rclone設定からの取得（ローカル環境用）
    try:
        result = subprocess.run(
            ["rclone", "config", "dump", "--config", RCLONE_CONFIG],
            capture_output=True, text=True,
        )
        config = json.loads(result.stdout)
        token_data = json.loads(config["manus_google_drive"]["token"])
        return token_data["access_token"]
    except Exception as e:
        logger.error(f"アクセストークン取得エラー: {e}")
        return ""


def download_spreadsheet() -> openpyxl.Workbook:
    token = get_access_token()
    url = (
        f"https://www.googleapis.com/drive/v3/files/{SPREADSHEET_ID}/export"
        f"?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content))


def upload_spreadsheet(wb: openpyxl.Workbook) -> None:
    token = get_access_token()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    url = (
        f"https://www.googleapis.com/upload/drive/v3/files/{SPREADSHEET_ID}"
        f"?uploadType=media&convert=true"
    )
    requests.patch(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=buf.read(),
    )


def get_or_create_sheet(wb: openpyxl.Workbook, sheet_name: str, headers: list) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    else:
        ws = wb[sheet_name]
    return ws


# ── 法人一覧取得 ───────────────────────────────────────────────────────────
def get_hojin_list() -> list[dict]:
    """スプレッドシートの法人一覧シートから法人データを取得する"""
    try:
        wb = download_spreadsheet()
        if "法人一覧シート" not in wb.sheetnames:
            return []
        ws = wb["法人一覧シート"]
        hojin_list = []
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if row[1]:  # B列: 法人名
                hojin = {
                    "name": str(row[1]),
                    "zip": str(row[2]) if row[2] and not str(row[2]).startswith("=") else "",
                    "pref": str(row[3]) if row[3] else "",
                    "city": str(row[4]) if row[4] else "",
                    "addr": str(row[5]) if row[5] else "",
                    "tel": str(row[6]) if row[6] else "",
                    "email": str(row[7]) if row[7] else "",
                    "tantousha": str(row[9]) if row[9] else "",
                }
                hojin_list.append(hojin)
        return hojin_list
    except Exception as e:
        logger.error(f"法人一覧取得エラー: {e}")
        return []


# ── /start コマンド ────────────────────────────────────────────────────────
WELCOME_MESSAGE = """スマホ1台の契約につき10,000円の報酬をお受け取りいただけます。法人契約のため、1日あたり5〜7万円前後の報酬が見込めます。

今なら登録完了で、現金8,000円プレゼント。

代行登録はコチら👇"""


# お仕事の流れ画像パス
OSHIGOTO_FLOW_IMAGE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "oshigoto_flow.jpg")


def _is_meishi_allowed(update: Update) -> bool:
    """名刺作成機能の利用が許可されているユーザーか判定する"""
    user = update.effective_user
    if not user:
        return False
    username = (user.username or "").lower()
    return username in MEISHI_ALLOWED_USERS


def _make_reply_keyboard() -> ReplyKeyboardMarkup:
    """常時表示するキーボードメニューを生成する"""
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton("🏦 支払い依頼"), KeyboardButton("📝 稼働データ入力")],
            [KeyboardButton("📋 メニューを表示")],
        ],
        resize_keyboard=True,
        persistent=True,
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # ウェルカムメッセージ：代行登録ボタン付き（常時表示キーボードメニューも同時に設置）
    welcome_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 代行登録フォーム", callback_data="menu_register")],
    ])
    await update.message.reply_text(
        WELCOME_MESSAGE,
        reply_markup=welcome_markup,
    )

    # お仕事の流れインフォグラフィック画像を送信
    try:
        with open(OSHIGOTO_FLOW_IMAGE, "rb") as photo:
            await update.message.reply_photo(photo=photo)
    except Exception as e:
        logger.error(f"お仕事の流れ画像の送信に失敗: {e}")

    # スタッフ用インラインボタンメニュー（常時表示キーボードも同時に設置）
    inline_keyboard = []
    if _is_meishi_allowed(update):
        inline_keyboard.append([InlineKeyboardButton("🪦 名刺の自動作成", callback_data="menu_meishi")])
    inline_keyboard.append([InlineKeyboardButton("🏦 支払い依頼フォーム", callback_data="menu_transfer")])
    inline_keyboard.append([InlineKeyboardButton("📝 稼働データ入力フォーム", callback_data="menu_report")])
    inline_markup = InlineKeyboardMarkup(inline_keyboard)
    await update.message.reply_text(
        "↓ スタッフ用メニュー：",
        reply_markup=InlineKeyboardMarkup(inline_keyboard),
    )
    # 常時表示キーボードを設置
    await update.message.reply_text(
        "↓ ボタンからいつでも操作できます。",
        reply_markup=_make_reply_keyboard(),
    )


async def menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """ConversationHandlerに属さないメニューコールバック用（フォールバック）"""
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "menu_meishi":
        # 許可ユーザー以外は拒否
        if not _is_meishi_allowed(update):
            await query.answer("この機能は利用できません。", show_alert=True)
            return
        return await start_meishi(update, context)
    elif data == "menu_transfer":
        return await start_transfer(update, context)
    elif data == "menu_report":
        return await start_report(update, context)
    elif data == "menu_register":
        return await start_register(update, context)


# ═══════════════════════════════════════════════════════════════════════════
# 機能1: 名刺の自動作成
# ═══════════════════════════════════════════════════════════════════════════

MEISHI_PAGE_SIZE = 8  # 1ページあたりの法人数


async def start_meishi(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    # 許可ユーザー以外は実行不可
    if not _is_meishi_allowed(update):
        await query.message.reply_text("❌ この機能は利用できません。")
        return ConversationHandler.END

    msg = await query.message.reply_text("⏳ 法人一覧を読み込み中...")
    hojin_list = get_hojin_list()

    if not hojin_list:
        await msg.edit_text("❌ 法人一覧の取得に失敗しました。")
        return ConversationHandler.END

    context.user_data["hojin_list"] = list(reversed(hojin_list))  # 新しい法人が先頭に来るよう逆順
    context.user_data["meishi_page"] = 0
    await msg.delete()
    return await show_meishi_page(update, context, query.message)


async def show_meishi_page(update: Update, context: ContextTypes.DEFAULT_TYPE, message=None) -> int:
    hojin_list = context.user_data.get("hojin_list", [])
    page = context.user_data.get("meishi_page", 0)
    total = len(hojin_list)
    total_pages = (total + MEISHI_PAGE_SIZE - 1) // MEISHI_PAGE_SIZE

    start_idx = page * MEISHI_PAGE_SIZE
    end_idx = min(start_idx + MEISHI_PAGE_SIZE, total)
    page_items = hojin_list[start_idx:end_idx]

    keyboard = []
    for i, h in enumerate(page_items):
        idx = start_idx + i
        # 法人名を短縮表示（長すぎる場合）
        display_name = h["name"]
        if len(display_name) > 20:
            display_name = display_name[:19] + "…"
        keyboard.append([InlineKeyboardButton(display_name, callback_data=f"meishi_select_{idx}")])

    # ページナビゲーション
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton("◀ 前へ", callback_data=f"meishi_page_{page - 1}"))
    if page < total_pages - 1:
        nav_row.append(InlineKeyboardButton("次へ ▶", callback_data=f"meishi_page_{page + 1}"))
    if nav_row:
        keyboard.append(nav_row)
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="meishi_cancel")])

    reply_markup = InlineKeyboardMarkup(keyboard)
    text = f"🪪 名刺を作成する法人を選択してください\n（{page + 1}/{total_pages}ページ、全{total}件）"

    if message:
        await message.reply_text(text, reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.message.edit_text(text, reply_markup=reply_markup)

    return MEISHI_SELECT


async def meishi_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "meishi_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    if data.startswith("meishi_page_"):
        page = int(data.split("_")[-1])
        context.user_data["meishi_page"] = page
        return await show_meishi_page(update, context)

    if data.startswith("meishi_select_"):
        idx = int(data.split("_")[-1])
        hojin_list = context.user_data.get("hojin_list", [])
        if idx >= len(hojin_list):
            await query.message.edit_text("❌ エラーが発生しました。")
            return ConversationHandler.END

        hojin = hojin_list[idx]
        await query.message.edit_text(f"⏳ 「{hojin['name']}」の名刺を生成中...")

        try:
            # 住所を組み立て
            address_parts = [hojin["pref"], hojin["city"], hojin["addr"]]
            address = "".join(p for p in address_parts if p)

            # 名刺PNG生成
            png_data = create_business_card(
                hojin_name=hojin["name"],
                tel=hojin["tel"],
                address=address,
                tantousha=hojin["tantousha"],
            )

            # 送信
            from meishi_generator import generate_email
            email = generate_email(hojin["name"])

            caption = (
                f"🪪 **{hojin['name']}**\n"
                f"担当者: {hojin['tantousha'] or '未設定'}\n"
                f"TEL: {hojin['tel'] or '未設定'}\n"
                f"MAIL: {email}\n"
                f"住所: {address or '未設定'}"
            )

            await query.message.reply_photo(
                photo=io.BytesIO(png_data),
                caption=caption,
                parse_mode="Markdown",
            )
            await query.message.edit_text("✅ 名刺を生成しました。")

        except Exception as e:
            logger.error(f"名刺生成エラー: {e}")
            await query.message.edit_text(f"❌ 名刺の生成に失敗しました: {e}")

        return ConversationHandler.END

    return MEISHI_SELECT


# ═══════════════════════════════════════════════════════════════════════════
# 機能2: 支払い依頼フォーム
# ═══════════════════════════════════════════════════════════════════════════

async def start_transfer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["transfer"] = {}
    if update.callback_query:
        await update.callback_query.message.reply_text(
            "🏦 **支払い依頼フォーム**\n\nお名前を入力してください：",
            parse_mode="Markdown",
        )
    else:
        await update.message.reply_text(
            "🏦 **支払い依頼フォーム**\n\nお名前を入力してください：",
            parse_mode="Markdown",
        )
    return TRANSFER_NAME


async def transfer_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["transfer"]["name"] = update.message.text.strip()

    # 主要銀行ボタン表示
    keyboard = []
    row = []
    for i, bank in enumerate(MAJOR_BANKS):
        row.append(InlineKeyboardButton(bank, callback_data=f"bank_{bank}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("🔍 その他（検索）", callback_data="bank_search")])
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="transfer_cancel")])

    await update.message.reply_text(
        "銀行名を選択してください：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return TRANSFER_BANK


async def transfer_bank_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "transfer_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    if data == "bank_search":
        await query.message.edit_text("🔍 銀行名の一部を入力してください（例：「三菱」「信金」など）：")
        return TRANSFER_BANK_SEARCH

    if data.startswith("bank_"):
        bank_name = data[5:]
        context.user_data["transfer"]["bank"] = bank_name
        await query.message.edit_text(f"✅ 銀行名：{bank_name}\n\n支店名を入力してください：")
        return TRANSFER_BRANCH

    return TRANSFER_BANK


async def transfer_bank_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    keyword = update.message.text.strip()
    results = search_banks(keyword)

    if not results:
        await update.message.reply_text(
            f"「{keyword}」に一致する銀行が見つかりませんでした。\n別のキーワードを入力してください："
        )
        return TRANSFER_BANK_SEARCH

    keyboard = []
    for bank in results[:10]:
        keyboard.append([InlineKeyboardButton(bank, callback_data=f"bank_{bank}")])
    keyboard.append([InlineKeyboardButton("🔍 再検索", callback_data="bank_search")])
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="transfer_cancel")])

    await update.message.reply_text(
        f"「{keyword}」の検索結果：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return TRANSFER_BANK


async def transfer_branch(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["transfer"]["branch"] = update.message.text.strip()
    # 口座種別は「普通」固定（ユーザーに選択させない）
    context.user_data["transfer"]["type"] = "普通"
    await update.message.reply_text("口座番号を入力してください：")
    return TRANSFER_ACCOUNT


async def transfer_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["transfer"]["account"] = update.message.text.strip()
    await update.message.reply_text("振込金額を入力してください（例：50000）：")
    return TRANSFER_AMOUNT


async def transfer_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    amount_text = update.message.text.strip().replace(",", "").replace("円", "")
    try:
        amount = int(amount_text)
        context.user_data["transfer"]["amount"] = amount
    except ValueError:
        await update.message.reply_text("❌ 金額は数字で入力してください：")
        return TRANSFER_AMOUNT

    t = context.user_data["transfer"]
    confirm_text = (
        f"📋 **振込依頼の確認**\n\n"
        f"お名前：{t['name']}\n"
        f"銀行名：{t['bank']}\n"
        f"支店名：{t['branch']}\n"
        f"口座種別：{t['type']}\n"
        f"口座番号：{t['account']}\n"
        f"振込金額：¥{amount:,}\n\n"
        f"この内容で送信しますか？"
    )
    keyboard = [
        [
            InlineKeyboardButton("✅ 送信", callback_data="transfer_submit"),
            InlineKeyboardButton("❌ キャンセル", callback_data="transfer_cancel"),
        ]
    ]
    await update.message.reply_text(
        confirm_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown",
    )
    return TRANSFER_CONFIRM


async def transfer_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "transfer_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    t = context.user_data["transfer"]
    try:
        wb = download_spreadsheet()
        ws = get_or_create_sheet(
            wb, "支払い依頼",
            ["タイムスタンプ", "お名前", "銀行名", "支店名", "口座種別", "口座番号", "振込金額"]
        )
        now = datetime.now(JST).strftime("%Y/%m/%d %H:%M:%S")
        ws.append([now, t["name"], t["bank"], t["branch"], t["type"], t["account"], t["amount"]])
        upload_spreadsheet(wb)
        await query.message.edit_text("✅ 振込依頼を送信しました。")

        # ── 振込依頼内容をグループに自動転送 ──────────────────────────────
        notify_text = (
            f"💰 **振込依頼が届きました**\n\n"
            f"受付日時：{now}\n"
            f"お名前：{t['name']}\n"
            f"銀行名：{t['bank']}\n"
            f"支店名：{t['branch']}\n"
            f"口座種別：{t['type']}\n"
            f"口座番号：{t['account']}\n"
            f"振込金額：¥{t['amount']:,}"
        )
        try:
            await context.bot.send_message(
                chat_id=TRANSFER_NOTIFY_GROUP_ID,
                text=notify_text,
                parse_mode="Markdown",
            )
            logger.info(f"振込依頼をグループ {TRANSFER_NOTIFY_GROUP_ID} に転送しました")
        except Exception as notify_err:
            logger.error(f"グループへの転送に失敗しました: {notify_err}")

    except Exception as e:
        logger.error(f"振込依頼送信エラー: {e}")
        await query.message.edit_text(f"❌ 送信に失敗しました: {e}")

    return ConversationHandler.END


# ═══════════════════════════════════════════════════════════════════════════
# 機能3: 稼働データ入力フォーム
# ═══════════════════════════════════════════════════════════════════════════

async def start_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["report"] = {"iphones": []}
    if update.callback_query:
        await update.callback_query.message.reply_text(
            "📝 **稼働データ入力フォーム**\n\n稼働者名を入力してください：",
            parse_mode="Markdown",
        )
    else:
        await update.message.reply_text(
            "📝 **稼働データ入力フォーム**\n\n稼働者名を入力してください：",
            parse_mode="Markdown",
        )
    return REPORT_NAME


async def report_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["report"]["name"] = update.message.text.strip()

    keyboard = [
        [
            InlineKeyboardButton("1店舗", callback_data="shop_1"),
            InlineKeyboardButton("2店舗", callback_data="shop_2"),
        ],
        [
            InlineKeyboardButton("3店舗", callback_data="shop_3"),
            InlineKeyboardButton("4店舗", callback_data="shop_4"),
        ],
    ]
    await update.message.reply_text(
        "稼働店舗数を選択してください：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return REPORT_SHOP


async def report_shop_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    shops = query.data.split("_")[1]
    context.user_data["report"]["shops"] = shops

    today = datetime.now(JST).strftime("%Y/%m/%d")
    keyboard = [
        [InlineKeyboardButton(f"今日（{today}）", callback_data="date_today")],
        [InlineKeyboardButton("日付を入力する", callback_data="date_input")],
    ]
    await query.message.edit_text(
        f"稼働店舗数：{shops}店舗\n\n稼働日を選択してください：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return REPORT_DATE


async def report_date_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "date_today":
        today = datetime.now(JST).strftime("%Y/%m/%d")
        context.user_data["report"]["date"] = today
        return await show_model_selection(query.message, context)
    else:
        await query.message.edit_text("稼働日を入力してください（例：2025/03/15）：")
        return REPORT_DATE_INPUT


async def report_date_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    date_text = update.message.text.strip()
    context.user_data["report"]["date"] = date_text
    return await show_model_selection(update.message, context)


async def show_model_selection(message, context: ContextTypes.DEFAULT_TYPE) -> int:
    iphones = context.user_data["report"].get("iphones", [])
    iphone_summary = ""
    if iphones:
        iphone_summary = "\n".join([f"  • {ip['model']} {ip['capacity']} × {ip['qty']}台" for ip in iphones])
        iphone_summary = f"\n\n**登録済み機種：**\n{iphone_summary}"

    keyboard = [
        [
            InlineKeyboardButton("iPhone 16", callback_data="model_iPhone16"),
            InlineKeyboardButton("iPhone 16e", callback_data="model_iPhone16e"),
        ],
    ]
    if iphones:
        keyboard.append([InlineKeyboardButton("✅ 入力完了", callback_data="model_done")])
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="report_cancel")])

    await message.reply_text(
        f"iPhone機種を選択してください：{iphone_summary}",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown",
    )
    return REPORT_MODEL


async def report_model_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "report_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    if data == "model_done":
        return await show_report_confirm(query.message, context)

    model = data.split("_")[1]  # "iPhone16" or "iPhone16e"
    context.user_data["report"]["current_model"] = model

    if model == "iPhone16":
        capacities = ["128GB", "256GB", "512GB", "1TB"]
    else:
        capacities = ["128GB", "256GB", "512GB"]

    keyboard = [[InlineKeyboardButton(c, callback_data=f"cap_{c}")] for c in capacities]
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="report_cancel")])

    display_model = "iPhone 16" if model == "iPhone16" else "iPhone 16e"
    await query.message.edit_text(
        f"{display_model}の容量を選択してください：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return REPORT_CAPACITY


async def report_capacity_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "report_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    capacity = data.split("_")[1]
    context.user_data["report"]["current_capacity"] = capacity
    model = context.user_data["report"]["current_model"]
    display_model = "iPhone 16" if model == "iPhone16" else "iPhone 16e"

    await query.message.edit_text(
        f"{display_model} {capacity}\n\n台数を入力してください（例：3）："
    )
    return REPORT_QUANTITY


async def report_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    qty_text = update.message.text.strip()
    try:
        qty = int(qty_text)
        if qty <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ 台数は1以上の整数で入力してください：")
        return REPORT_QUANTITY

    model = context.user_data["report"]["current_model"]
    capacity = context.user_data["report"]["current_capacity"]
    display_model = "iPhone 16" if model == "iPhone16" else "iPhone 16e"

    context.user_data["report"]["iphones"].append({
        "model": display_model,
        "capacity": capacity,
        "qty": qty,
    })

    keyboard = [
        [
            InlineKeyboardButton("➕ 別の機種も追加", callback_data="model_add"),
            InlineKeyboardButton("✅ 入力完了", callback_data="model_done"),
        ],
        [InlineKeyboardButton("❌ キャンセル", callback_data="report_cancel")],
    ]
    iphones = context.user_data["report"]["iphones"]
    summary = "\n".join([f"  • {ip['model']} {ip['capacity']} × {ip['qty']}台" for ip in iphones])

    await update.message.reply_text(
        f"✅ 追加しました：{display_model} {capacity} × {qty}台\n\n**現在の登録：**\n{summary}\n\n続けますか？",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown",
    )
    return REPORT_ADD_MORE


async def report_add_more_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "report_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    if data == "model_done":
        return await show_report_confirm(query.message, context)

    if data == "model_add":
        return await show_model_selection(query.message, context)

    return REPORT_ADD_MORE


async def show_report_confirm(message, context: ContextTypes.DEFAULT_TYPE) -> int:
    r = context.user_data["report"]
    iphones = r.get("iphones", [])
    iphone_text = "\n".join([f"  • {ip['model']} {ip['capacity']} × {ip['qty']}台" for ip in iphones])

    confirm_text = (
        f"📋 **稼働データの確認**\n\n"
        f"稼働者名：{r['name']}\n"
        f"稼働店舗数：{r['shops']}店舗\n"
        f"稼働日：{r['date']}\n"
        f"iPhone詳細：\n{iphone_text}\n\n"
        f"この内容で送信しますか？"
    )
    keyboard = [
        [
            InlineKeyboardButton("✅ 送信", callback_data="report_submit"),
            InlineKeyboardButton("❌ キャンセル", callback_data="report_cancel"),
        ]
    ]
    await message.reply_text(
        confirm_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown",
    )
    return REPORT_CONFIRM


async def report_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "report_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    r = context.user_data["report"]
    iphones = r.get("iphones", [])
    iphone_detail = " / ".join([f"{ip['model']} {ip['capacity']} × {ip['qty']}台" for ip in iphones])

    try:
        wb = download_spreadsheet()
        ws = get_or_create_sheet(
            wb, "稼働データ",
            ["タイムスタンプ", "稼働者名", "稼働店舗数", "稼働日", "iPhone詳細"]
        )
        now = datetime.now(JST).strftime("%Y/%m/%d %H:%M:%S")
        ws.append([now, r["name"], r["shops"], r["date"], iphone_detail])
        upload_spreadsheet(wb)
        await query.message.edit_text("✅ 稼働データを送信しました。")
    except Exception as e:
        logger.error(f"稼働報告送信エラー: {e}")
        await query.message.edit_text(f"❌ 送信に失敗しました: {e}")

    return ConversationHandler.END


# ═══════════════════════════════════════════════════════════════════════════
# 機能4: 代行登録フォーム
# ═══════════════════════════════════════════════════════════════════════════

async def start_register(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """代行登録フォームの開始"""
    context.user_data["register"] = {}
    msg_text = (
        "📝 **代行登録フォーム**\n\n"
        "お名前（氏名）を入力してください："
    )
    if update.callback_query:
        await update.callback_query.message.reply_text(msg_text, parse_mode="Markdown")
    else:
        await update.message.reply_text(msg_text, parse_mode="Markdown")
    return REG_NAME


async def reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["register"]["name"] = update.message.text.strip()
    await update.message.reply_text("電話番号を入力してください：")
    return REG_PHONE


async def reg_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["register"]["phone"] = update.message.text.strip()
    await update.message.reply_text("住所（都道府県から番地まで）を入力してください：")
    return REG_ADDRESS


async def reg_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["register"]["address"] = update.message.text.strip()
    # 銀行名入力へ
    keyboard = []
    row = []
    for i, bank in enumerate(MAJOR_BANKS):
        row.append(InlineKeyboardButton(bank, callback_data=f"reg_bank_{bank}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("🔍 銀行を検索", callback_data="reg_bank_search")])
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="reg_cancel")])
    await update.message.reply_text(
        "口座の銀行を選択してください：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return REG_BANK


async def reg_bank_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "reg_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    if data == "reg_bank_search":
        await query.message.edit_text("🔍 銀行名の一部を入力してください（例：「三菱」「信金」など）：")
        return REG_BANK_SEARCH

    if data.startswith("reg_bank_"):
        bank_name = data[9:]
        context.user_data["register"]["bank"] = bank_name
        await query.message.edit_text(f"✅ 銀行名：{bank_name}\n\n支店名を入力してください：")
        return REG_BRANCH

    return REG_BANK


async def reg_bank_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    keyword = update.message.text.strip()
    results = search_banks(keyword)

    if not results:
        await update.message.reply_text(
            f"「{keyword}」に一致する銀行が見つかりませんでした。\n別のキーワードを入力してください："
        )
        return REG_BANK_SEARCH

    keyboard = []
    for bank in results[:10]:
        keyboard.append([InlineKeyboardButton(bank, callback_data=f"reg_bank_{bank}")])
    keyboard.append([InlineKeyboardButton("🔍 再検索", callback_data="reg_bank_search")])
    keyboard.append([InlineKeyboardButton("❌ キャンセル", callback_data="reg_cancel")])
    await update.message.reply_text(
        f"「{keyword}」の検索結果：",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return REG_BANK


async def reg_branch(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["register"]["branch"] = update.message.text.strip()
    context.user_data["register"]["account_type"] = "普通"  # 口座種別は普通固定
    await update.message.reply_text("口座番号を入力してください：")
    return REG_ACCOUNT


async def reg_account(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["register"]["account"] = update.message.text.strip()
    r = context.user_data["register"]
    confirm_text = (
        f"📋 **代行登録内容の確認**\n\n"
        f"お名前：{r['name']}\n"
        f"電話番号：{r['phone']}\n"
        f"住所：{r['address']}\n"
        f"銀行名：{r['bank']}\n"
        f"支店名：{r['branch']}\n"
        f"口座種別：{r['account_type']}\n"
        f"口座番号：{r['account']}\n\n"
        f"この内容で登録しますか？"
    )
    keyboard = [[
        InlineKeyboardButton("✅ 登録する", callback_data="reg_submit"),
        InlineKeyboardButton("❌ キャンセル", callback_data="reg_cancel"),
    ]]
    await update.message.reply_text(
        confirm_text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown",
    )
    return REG_CONFIRM


async def reg_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "reg_cancel":
        await query.message.edit_text("❌ キャンセルしました。")
        return ConversationHandler.END

    r = context.user_data["register"]
    user = query.from_user
    tg_username = f"@{user.username}" if user.username else str(user.id)

    try:
        wb = download_spreadsheet()
        ws = get_or_create_sheet(
            wb, "代行登録",
            ["タイムスタンプ", "TGユーザー名", "お名前", "電話番号", "住所", "銀行名", "支店名", "口座種別", "口座番号"]
        )
        now = datetime.now(JST).strftime("%Y/%m/%d %H:%M:%S")
        ws.append([
            now,
            tg_username,
            r["name"],
            r["phone"],
            r["address"],
            r["bank"],
            r["branch"],
            r["account_type"],
            r["account"],
        ])
        upload_spreadsheet(wb)
        await query.message.edit_text(
            "✅ 登録が完了しました！\n担当者からご連絡いたします。しばらくお待ちください。"
        )
    except Exception as e:
        logger.error(f"代行登録保存エラー: {e}")
        await query.message.edit_text(f"❌ 登録に失敗しました: {e}")

    return ConversationHandler.END


# ═══════════════════════════════════════════════════════════════════════════
# メイン
# ═══════════════════════════════════════════════════════════════════════════

async def post_init(application: Application) -> None:
    """起動時にset_my_commandsでコマンドメニューを登録する"""
    commands = [
        BotCommand("start", "メニューを表示する"),
        BotCommand("menu", "メニューを表示する"),
        BotCommand("transfer", "支払い依頼フォーム"),
        BotCommand("report", "稼働データ入力フォーム"),
    ]
    await application.bot.set_my_commands(commands)
    logger.info("Bot commands registered.")


async def show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/menuコマンドやキーボードボタンからインラインメニューを表示する"""
    inline_keyboard = []
    if _is_meishi_allowed(update):
        inline_keyboard.append([InlineKeyboardButton("🪦 名刺の自動作成", callback_data="menu_meishi")])
    inline_keyboard.append([InlineKeyboardButton("🏦 支払い依頼フォーム", callback_data="menu_transfer")])
    inline_keyboard.append([InlineKeyboardButton("📝 稼働データ入力フォーム", callback_data="menu_report")])
    inline_markup = InlineKeyboardMarkup(inline_keyboard)
    msg = update.message or (update.callback_query and update.callback_query.message)
    if msg:
        await msg.reply_text(
            "↓ 以下のボタンから選択してください：",
            reply_markup=inline_markup,
        )


async def _keyboard_button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """常時表示キーボードのボタンテキストに応じて各機能を起動する"""
    text = update.message.text
    if text == "🏦 支払い依頼":
        await start_transfer(update, context)
    elif text == "📝 稼働データ入力":
        await start_report(update, context)
    else:  # "📋 メニューを表示"
        await show_menu(update, context)


def main() -> None:
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    # /start コマンド
    app.add_handler(CommandHandler("start", start))

    # /menu コマンド
    app.add_handler(CommandHandler("menu", show_menu))

    # キーボード「📋 メニューを表示」ボタンのハンドラ（インラインメニューを表示）
    app.add_handler(MessageHandler(
        filters.TEXT & filters.Regex(r"^📋 メニューを表示$"),
        show_menu,
    ))

    # 名刺作成 ConversationHandler
    meishi_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(start_meishi, pattern="^menu_meishi$")],
        states={
            MEISHI_SELECT: [CallbackQueryHandler(meishi_callback, pattern="^meishi_")],
        },
        fallbacks=[
            CommandHandler("start", start),
            CallbackQueryHandler(menu_callback, pattern="^menu_"),
        ],
        per_user=True,
        per_chat=True,
        allow_reentry=True,
    )
    app.add_handler(meishi_conv)

    # 支払い依頼 ConversationHandler
    transfer_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(start_transfer, pattern="^menu_transfer$"),
            CommandHandler("transfer", start_transfer),
            MessageHandler(filters.TEXT & filters.Regex(r"^🏦 支払い依頼$"), start_transfer),
        ],
        states={
            TRANSFER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_name)],
            TRANSFER_BANK: [CallbackQueryHandler(transfer_bank_callback, pattern="^(bank_|transfer_cancel)")],
            TRANSFER_BANK_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_bank_search),
                CallbackQueryHandler(transfer_bank_callback, pattern="^(bank_|transfer_cancel)"),
            ],
            TRANSFER_BRANCH: [MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_branch)],
            TRANSFER_ACCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_account)],
            TRANSFER_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, transfer_amount)],
            TRANSFER_CONFIRM: [CallbackQueryHandler(transfer_confirm_callback, pattern="^transfer_")],
        },
        fallbacks=[
            CommandHandler("start", start),
            CallbackQueryHandler(menu_callback, pattern="^menu_"),
        ],
        per_user=True,
        per_chat=True,
        allow_reentry=True,
    )
    app.add_handler(transfer_conv)

    # 稼働報告 ConversationHandler
    report_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(start_report, pattern="^menu_report$"),
            CommandHandler("report", start_report),
            MessageHandler(filters.TEXT & filters.Regex(r"^📝 稼働データ入力$"), start_report),
        ],
        states={
            REPORT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, report_name)],
            REPORT_SHOP: [CallbackQueryHandler(report_shop_callback, pattern="^shop_")],
            REPORT_DATE: [CallbackQueryHandler(report_date_callback, pattern="^date_")],
            REPORT_DATE_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, report_date_input)],
            REPORT_MODEL: [CallbackQueryHandler(report_model_callback, pattern="^(model_|report_cancel)")],
            REPORT_CAPACITY: [CallbackQueryHandler(report_capacity_callback, pattern="^(cap_|report_cancel)")],
            REPORT_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, report_quantity)],
            REPORT_ADD_MORE: [CallbackQueryHandler(report_add_more_callback, pattern="^(model_|report_cancel)")],
            REPORT_CONFIRM: [CallbackQueryHandler(report_confirm_callback, pattern="^report_")],
        },
        fallbacks=[
            CommandHandler("start", start),
            CallbackQueryHandler(menu_callback, pattern="^menu_"),
        ],
        per_user=True,
        per_chat=True,
        allow_reentry=True,
    )
    app.add_handler(report_conv)

    # 代行登録 ConversationHandler
    register_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(start_register, pattern="^menu_register$"),
            CommandHandler("register", start_register),
        ],
        states={
            REG_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)],
            REG_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_phone)],
            REG_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_address)],
            REG_BANK: [CallbackQueryHandler(reg_bank_callback, pattern="^(reg_bank_|reg_cancel)")],
            REG_BANK_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, reg_bank_search),
                CallbackQueryHandler(reg_bank_callback, pattern="^(reg_bank_|reg_cancel)"),
            ],
            REG_BRANCH: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_branch)],
            REG_ACCOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_account)],
            REG_CONFIRM: [CallbackQueryHandler(reg_confirm_callback, pattern="^reg_")],
        },
        fallbacks=[
            CommandHandler("start", start),
            CallbackQueryHandler(menu_callback, pattern="^menu_"),
        ],
        per_user=True,
        per_chat=True,
        allow_reentry=True,
    )
    app.add_handler(register_conv)

    # メニューコールバック（ConversationHandlerにマッチしない場合のフォールバック）
    app.add_handler(CallbackQueryHandler(menu_callback, pattern="^menu_"))

    logger.info("Bot started.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
